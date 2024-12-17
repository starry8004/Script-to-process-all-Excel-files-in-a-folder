import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import os
from datetime import datetime

class ExcelProcessor:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("엑셀 파일 일괄 처리")
        self.root.geometry("600x250")
        
        # UI 구성요소
        self.create_widgets()
        
    def create_widgets(self):
        # 파일 정보 표시
        self.file_label = tk.Label(self.root, text="현재 처리중인 파일: ")
        self.file_label.pack(pady=5)
        
        # 전체 진행상황 표시
        self.total_progress_label = tk.Label(self.root, text="전체 진행률:")
        self.total_progress_label.pack(pady=5)
        self.total_progress_bar = ttk.Progressbar(self.root, length=400, mode='determinate')
        self.total_progress_bar.pack(pady=5)
        
        # 현재 파일 진행상황 표시
        self.current_progress_label = tk.Label(self.root, text="현재 파일 진행률:")
        self.current_progress_label.pack(pady=5)
        self.current_progress_bar = ttk.Progressbar(self.root, length=400, mode='determinate')
        self.current_progress_bar.pack(pady=5)
        
        # 상세 상태 표시
        self.status_label = tk.Label(self.root, text="대기 중...")
        self.status_label.pack(pady=10)

    def process_single_file(self, input_file):
        try:
            # 파일명 표시 업데이트
            self.file_label.config(text=f"현재 처리중인 파일: {os.path.basename(input_file)}")
            self.root.update()
            
            # 엑셀 파일 읽기
            self.status_label.config(text="파일을 읽는 중...")
            self.root.update()
            wb = openpyxl.load_workbook(input_file)
            sheet = wb.active
            
            # 헤더 찾기
            self.status_label.config(text="'검색량' 열을 찾는 중...")
            self.root.update()
            header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            search_col_idx = None
            for idx, header in enumerate(header_row, 1):
                if header == '검색량':
                    search_col_idx = idx
                    break
            
            if search_col_idx is None:
                raise ValueError(f"'{input_file}'에서 '검색량' 열을 찾을 수 없습니다.")
            
            # 새 워크북 생성
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active
            
            # 헤더 복사
            for col in range(1, sheet.max_column + 1):
                new_sheet.cell(1, col, sheet.cell(1, col).value)
            
            # 전체 행 수 계산
            total_rows = sheet.max_row - 1  # 헤더 제외
            
            # 데이터 필터링
            new_row = 2
            filtered_count = 0
            
            for row in range(2, sheet.max_row + 1):
                # 현재 파일 진행률 업데이트
                progress = (row - 2) / total_rows * 100
                self.current_progress_bar['value'] = progress
                self.status_label.config(text=f"처리 중... ({row-1}/{total_rows} 행)")
                self.root.update()
                
                search_value = sheet.cell(row, search_col_idx).value
                try:
                    if isinstance(search_value, str):
                        search_value = int(search_value.replace(',', ''))
                    else:
                        search_value = int(search_value)
                        
                    if search_value >= 8000:
                        for col in range(1, sheet.max_column + 1):
                            new_sheet.cell(new_row, col, sheet.cell(row, col).value)
                        new_row += 1
                        filtered_count += 1
                except (ValueError, TypeError):
                    continue
            
            # 저장
            self.status_label.config(text="파일 저장 중...")
            self.root.update()
            
            # 현재 날짜와 시간을 포함한 파일명 생성
            current_time = datetime.now().strftime('%Y-%m-%d_%H-%M')
            file_dir = os.path.dirname(input_file)
            file_name = os.path.splitext(os.path.basename(input_file))[0]
            output_file = os.path.join(file_dir, f"{file_name}_8000del_{current_time}.xlsx")
            new_wb.save(output_file)
            
            return {
                'status': 'success',
                'total_rows': total_rows,
                'filtered_count': filtered_count,
                'output_file': output_file
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'error_message': str(e)
            }

    def process_folder(self):
        # 폴더 선택 다이얼로그
        folder_path = filedialog.askdirectory(title="처리할 엑셀 파일이 있는 폴더를 선택하세요")
        
        if not folder_path:  # 폴더 선택 취소시
            self.root.destroy()
            return
        
        # 폴더 내 모든 Excel 파일 찾기
        excel_files = [f for f in os.listdir(folder_path) 
                      if f.endswith('.xlsx') and not f.endswith('_8000del.xlsx')]
        
        if not excel_files:
            messagebox.showwarning("경고", "선택한 폴더에 Excel 파일이 없습니다.")
            self.root.destroy()
            return
        
        # 처리 결과 저장
        results = {
            'success': 0,
            'error': 0,
            'details': []
        }
        
        # 전체 파일 처리
        for idx, file_name in enumerate(excel_files):
            # 전체 진행률 업데이트
            total_progress = (idx / len(excel_files)) * 100
            self.total_progress_bar['value'] = total_progress
            self.total_progress_label.config(text=f"전체 진행률: {idx + 1}/{len(excel_files)} 파일")
            
            # 현재 파일 진행률 초기화
            self.current_progress_bar['value'] = 0
            
            # 파일 처리
            input_file = os.path.join(folder_path, file_name)
            result = self.process_single_file(input_file)
            
            if result['status'] == 'success':
                results['success'] += 1
                results['details'].append({
                    'file_name': file_name,
                    'status': 'success',
                    'total_rows': result['total_rows'],
                    'filtered_count': result['filtered_count'],
                    'output_file': result['output_file']
                })
            else:
                results['error'] += 1
                results['details'].append({
                    'file_name': file_name,
                    'status': 'error',
                    'error_message': result['error_message']
                })
        
        # 최종 결과 표시
        self.show_results(results)
        
    def show_results(self, results):
        result_message = f"처리가 완료되었습니다!\n\n"
        result_message += f"성공: {results['success']} 파일\n"
        result_message += f"실패: {results['error']} 파일\n\n"
        result_message += "상세 결과:\n"
        
        for detail in results['details']:
            result_message += f"\n● {detail['file_name']}\n"
            if detail['status'] == 'success':
                result_message += f"  - 전체 데이터: {detail['total_rows']}행\n"
                result_message += f"  - 필터링된 데이터: {detail['filtered_count']}행\n"
                result_message += f"  - 저장 위치: {detail['output_file']}\n"
            else:
                result_message += f"  - 오류: {detail['error_message']}\n"
        
        messagebox.showinfo("처리 결과", result_message)
        self.root.destroy()

    def run(self):
        self.process_folder()
        self.root.mainloop()

if __name__ == "__main__":
    processor = ExcelProcessor()
    processor.run()